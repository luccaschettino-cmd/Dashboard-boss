from flask import Flask, render_template, jsonify, send_file, request
import requests
import os
import json
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from dotenv import load_dotenv
from collections import defaultdict
import re
import threading
import psycopg2
import psycopg2.extras

load_dotenv()

app = Flask(__name__)

TOKEN = os.getenv("EAD_TOKEN")
BASE_URL = "https://bosstreinamentos.com/api/1"
HEADERS = {"x-auth-token": TOKEN}
DATABASE_URL = os.getenv("DATABASE_URL")


VALIDADE_NR = [
    ("nr 33", 365), ("nr 20 intermedi", 730), ("nr 20 avanc", 730),
    ("nr 20 basic", 1095), ("nr 20", 1095), ("nr 35", 730), ("nr 10", 730),
    ("nr 12", 730), ("nr 06", 730), ("nr 05", 730), ("nr 18", 730),
    ("nr 37", 730), ("nr 34", 730), ("direção defensiva", 1095),
    ("direcao defensiva", 1095), ("primeiros socorros", 730), ("cbasi", 730),
]
VALIDADE_PADRAO = 730

UF_SIGLAS = {
    "ACRE": "AC", "ALAGOAS": "AL", "AMAPÁ": "AP", "AMAZONAS": "AM",
    "BAHIA": "BA", "CEARÁ": "CE", "DISTRITO FEDERAL": "DF",
    "ESPÍRITO SANTO": "ES", "GOIÁS": "GO", "MARANHÃO": "MA",
    "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS", "MINAS GERAIS": "MG",
    "PARÁ": "PA", "PARAÍBA": "PB", "PARANÁ": "PR", "PERNAMBUCO": "PE",
    "PIAUÍ": "PI", "RIO DE JANEIRO": "RJ", "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS", "RONDÔNIA": "RO", "RORAIMA": "RR",
    "SANTA CATARINA": "SC", "SÃO PAULO": "SP", "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


def get_conn():
    return psycopg2.connect(DATABASE_URL)


def validade_curso(titulo, regras=None):
    if not titulo:
        return VALIDADE_PADRAO
    t = titulo.lower()
    for padrao, dias in (regras or get_validades()):
        if padrao in t:
            return dias
    return VALIDADE_PADRAO


def normalizar_uf(uf_raw):
    if not uf_raw:
        return ""
    v = uf_raw.strip().upper()
    if len(v) == 2:
        return v
    return UF_SIGLAS.get(v, v[:2] if len(v) >= 2 else v)


sync_status = {"running": False, "progress": "", "last_sync": None, "total": 0, "done": 0}
_table_cache: dict = {}


def init_db():
    conn = get_conn()
    c = conn.cursor()
    # Tabelas principais — id é TEXT para suportar IDs da API e fallback "o{offset}_{i}"
    for table in ("enrollments", "certificates", "students"):
        c.execute(f"""
            CREATE TABLE IF NOT EXISTS {table} (
                id TEXT PRIMARY KEY,
                data TEXT NOT NULL,
                synced_at TEXT NOT NULL
            )
        """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS emails_enviados (
            id SERIAL PRIMARY KEY,
            aluno_id BIGINT NOT NULL,
            curso_id BIGINT NOT NULL,
            faixa INTEGER NOT NULL,
            data_envio TEXT NOT NULL
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_emails_aluno_curso ON emails_enviados(aluno_id, curso_id, faixa)")
    c.execute("""
        CREATE TABLE IF NOT EXISTS validades (
            id SERIAL PRIMARY KEY,
            padrao TEXT NOT NULL,
            dias INTEGER NOT NULL,
            ordem INTEGER NOT NULL DEFAULT 0
        )
    """)
    # Seed validades padrão se tabela vazia
    c.execute("SELECT COUNT(*) FROM validades")
    if c.fetchone()[0] == 0:
        psycopg2.extras.execute_values(
            c,
            "INSERT INTO validades (padrao, dias, ordem) VALUES %s",
            [(padrao, dias, i) for i, (padrao, dias) in enumerate(VALIDADE_NR)]
        )
    conn.commit()
    c.close()
    conn.close()


init_db()


def get_validades():
    """Retorna lista de (padrao, dias) ordenada para uso em validade_curso()."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT padrao, dias FROM validades ORDER BY ordem")
    rows = c.fetchall()
    c.close()
    conn.close()
    return [(r[0], r[1]) for r in rows] if rows else VALIDADE_NR


def get_meta(key):
    try:
        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT value FROM meta WHERE key=%s", (key,))
        row = c.fetchone()
        c.close()
        conn.close()
        return row[0] if row else None
    except Exception as e:
        app.logger.warning("get_meta(%s) falhou: %s", key, e)
        return None


def count_table(table):
    try:
        conn = get_conn()
        c = conn.cursor()
        c.execute(f"SELECT COUNT(*) FROM {table}")
        n = c.fetchone()[0]
        c.close()
        conn.close()
        return n
    except Exception as e:
        app.logger.warning("count_table(%s) falhou: %s", table, e)
        return 0


def load_table(table):
    if table not in _table_cache:
        conn = get_conn()
        c = conn.cursor()
        c.execute(f"SELECT data FROM {table}")
        rows = c.fetchall()
        c.close()
        conn.close()
        _table_cache[table] = [json.loads(r[0]) for r in rows]
    return _table_cache[table]


def sync_endpoint(endpoint, table, label, id_field=None, data_minima="2023-01-01"):
    """Busca endpoint paginado e salva no PostgreSQL registros a partir de data_minima."""
    LIMIT = 200
    CAMPOS_DATA = {"enrollment": "cadastro", "certificate": "concluido", "progress": None, "student": None}
    campo_data = CAMPOS_DATA.get(endpoint)

    # Limpa a tabela antes de começar
    try:
        conn = get_conn()
        c = conn.cursor()
        c.execute(f"DELETE FROM {table}")
        conn.commit()
        c.close()
        conn.close()
    except Exception as e:
        app.logger.exception("sync_endpoint DELETE %s falhou", table)
        sync_status["progress"] = f"{label}: erro ao limpar tabela — {type(e).__name__}"
        return 0

    offset = 0
    total_salvos = 0
    total_lidos = 0

    while True:
        # Busca da API (sem conexão DB aberta)
        try:
            r = requests.get(f"{BASE_URL}/{endpoint}", headers=HEADERS,
                             params={"limit": LIMIT, "offset": offset}, timeout=30)
            if r.status_code != 200:
                app.logger.warning("sync_endpoint %s offset %d: HTTP %d", label, offset, r.status_code)
                break
            data = r.json()
            if not data:
                break
            items = data if isinstance(data, list) else (data.get("data") or data.get("results") or [])
            if not items:
                break
        except requests.RequestException as e:
            sync_status["progress"] = f"{label}: erro na requisição (offset {offset}) — {type(e).__name__}"
            app.logger.warning("sync_endpoint %s offset %d RequestException: %s", label, offset, e)
            break

        total_lidos += len(items)
        now_str = datetime.now().isoformat()
        batch = []
        for i, e in enumerate(items):
            if campo_data and data_minima:
                val = (e.get(campo_data) or "")[:10]
                if val and val < data_minima:
                    continue
            if id_field:
                row_id = e.get(id_field) or f"o{offset}_{i}"
            else:
                row_id = e.get("id") or e.get("matricula_id") or e.get("aluno_id") or f"o{offset}_{i}"
            batch.append((str(row_id), json.dumps(e), now_str))

        # Salva batch com nova conexão (evita timeout de conexão ociosa)
        if batch:
            try:
                conn = get_conn()
                c = conn.cursor()
                psycopg2.extras.execute_values(
                    c,
                    f"""INSERT INTO {table} (id, data, synced_at) VALUES %s
                        ON CONFLICT (id) DO UPDATE SET data=EXCLUDED.data, synced_at=EXCLUDED.synced_at""",
                    batch
                )
                conn.commit()
                c.close()
                conn.close()
                total_salvos += len(batch)
            except Exception as e:
                app.logger.exception("sync_endpoint %s offset %d: erro ao salvar batch", label, offset)
                sync_status["progress"] = f"{label} erro ao salvar offset {offset}: {type(e).__name__}"
                break

        sync_status["progress"] = f"{label}: {total_lidos} lidos, {total_salvos} salvos"
        sync_status["done"] = total_salvos
        if len(items) < LIMIT:
            break
        offset += LIMIT

    return total_salvos


def sync_endpoint_rapido(endpoint, table, label, paginas=10, id_field=None):
    """Busca apenas as últimas `paginas` páginas e faz upsert sem apagar o histórico."""
    LIMIT = 200
    total_no_banco = count_table(table)
    offset_inicial = max(0, total_no_banco - paginas * LIMIT)

    total_lidos = 0
    total_salvos = 0
    offset = offset_inicial

    while True:
        # Busca da API (sem conexão DB aberta)
        try:
            r = requests.get(f"{BASE_URL}/{endpoint}", headers=HEADERS,
                             params={"limit": LIMIT, "offset": offset}, timeout=30)
            if r.status_code != 200:
                break
            data = r.json()
            if not data:
                break
            items = data if isinstance(data, list) else (data.get("data") or data.get("results") or [])
            if not items:
                break
        except requests.RequestException as e:
            sync_status["progress"] = f"{label}: erro na requisição (offset {offset}) — {type(e).__name__}"
            break

        total_lidos += len(items)
        now_str = datetime.now().isoformat()
        batch = []
        for i, e in enumerate(items):
            if id_field:
                row_id = e.get(id_field) or f"o{offset}_{i}"
            else:
                row_id = e.get("id") or e.get("matricula_id") or e.get("aluno_id") or f"o{offset}_{i}"
            batch.append((str(row_id), json.dumps(e), now_str))

        # Salva batch com nova conexão
        try:
            conn = get_conn()
            c = conn.cursor()
            psycopg2.extras.execute_values(
                c,
                f"""INSERT INTO {table} (id, data, synced_at) VALUES %s
                    ON CONFLICT (id) DO UPDATE SET data=EXCLUDED.data, synced_at=EXCLUDED.synced_at""",
                batch
            )
            conn.commit()
            c.close()
            conn.close()
            total_salvos += len(items)
        except Exception as e:
            app.logger.exception("sync_endpoint_rapido %s offset %d: erro ao salvar batch", label, offset)
            sync_status["progress"] = f"{label} erro ao salvar offset {offset}: {type(e).__name__}"
            break

        sync_status["progress"] = f"{label} (rápido): {total_lidos} verificados, {total_salvos} atualizados"
        sync_status["done"] = total_salvos
        if len(items) < LIMIT:
            break
        offset += LIMIT

    return total_salvos


def _salvar_meta_sync(now_str, modo):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO meta (key, value) VALUES (%s, %s)
        ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value
    """, ("last_sync", now_str))
    c.execute("""
        INSERT INTO meta (key, value) VALUES (%s, %s)
        ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value
    """, ("last_sync_modo", modo))
    conn.commit()
    c.close()
    conn.close()


def do_sync(modo="completo"):
    global sync_status
    sync_status["running"] = True
    sync_status["progress"] = "Iniciando..."
    sync_status["done"] = 0
    try:
        if modo == "rapido":
            n_enr = sync_endpoint_rapido("enrollment", "enrollments", "Matrículas", paginas=10)
            n_cert = sync_endpoint_rapido("certificate", "certificates", "Certificados", paginas=10)
            n_stu = sync_endpoint_rapido("student", "students", "Alunos", paginas=3, id_field="aluno_id")
            msg = f"Rápido — {n_enr} matrículas, {n_cert} certificados, {n_stu} alunos."
        else:
            n_enr = sync_endpoint("enrollment", "enrollments", "Matrículas")
            n_cert = sync_endpoint("certificate", "certificates", "Certificados")
            n_stu = sync_endpoint("student", "students", "Alunos", id_field="aluno_id")
            msg = f"Completo — {n_enr} matrículas, {n_cert} certificados, {n_stu} alunos."

        now_str = datetime.now().isoformat()
        _salvar_meta_sync(now_str, modo)
        sync_status["last_sync"] = now_str
        sync_status["progress"] = msg
        _table_cache.clear()
    except Exception as e:
        sync_status["progress"] = f"Erro geral: {type(e).__name__}"
        app.logger.exception("do_sync falhou")
    finally:
        sync_status["running"] = False


def extract_empresa(grupo_nome):
    if not grupo_nome:
        return None
    parts = grupo_nome.split(" - ")
    if len(parts) >= 2:
        empresa = parts[1].strip()
        empresa = re.sub(r'\s+\d+$', '', empresa).strip()
        return empresa.upper() if empresa else None
    return grupo_nome.strip().upper()


def mapa_student():
    m = {}
    for s in load_table("students"):
        aid = s.get("aluno_id")
        if aid is not None:
            m[aid] = s
    return m


def mapa_email():
    m = {}
    for e in load_table("enrollments"):
        aid = e.get("aluno_id")
        email = e.get("aluno_email")
        if aid and email and aid not in m:
            m[aid] = email
    return m


def mapa_canal():
    m = {}
    for e in load_table("enrollments"):
        aid = e.get("aluno_id")
        if aid is None:
            continue
        canal = "B2B" if e.get("grupo_nome") else "B2C"
        if aid not in m or canal == "B2B":
            m[aid] = canal
    return m


def mapa_empresa():
    m = {}
    for e in load_table("enrollments"):
        aid = e.get("aluno_id")
        if aid and e.get("grupo_nome"):
            emp = extract_empresa(e["grupo_nome"])
            if emp:
                m[aid] = emp
    return m


def compute_geo(enrollments, students_map):
    b2c_ativos_ids = {
        e.get("aluno_id")
        for e in enrollments
        if e.get("status") == 1 and not e.get("grupo_nome") and e.get("aluno_id") is not None
    }
    contagem = defaultdict(int)
    for aid in b2c_ativos_ids:
        s = students_map.get(aid)
        if s:
            sigla = normalizar_uf(s.get("uf", ""))
            if sigla:
                contagem[sigla] += 1
    return sorted(contagem.items(), key=lambda x: x[1], reverse=True)


def anos_disponiveis(enrollments):
    anos = sorted({
        int((e.get("cadastro") or "")[:4])
        for e in enrollments
        if (e.get("cadastro") or "")[:4].isdigit()
    })
    return anos


def compute_vendas(enrollments, ano):
    now = datetime.now()
    ano_atual = now.year
    mes_atual = now.strftime("%Y-%m")
    ano_str = str(ano)

    b2b = [e for e in enrollments if e.get("grupo_nome")]
    b2c = [e for e in enrollments if not e.get("grupo_nome")]

    b2b_active = [e for e in b2b if e.get("status") == 1]
    b2c_active = [e for e in b2c if e.get("status") == 1]
    total_active = len(b2b_active) + len(b2c_active)

    empresa_count_ativas = defaultdict(int)
    for e in b2b_active:
        emp = extract_empresa(e.get("grupo_nome", ""))
        if emp:
            empresa_count_ativas[emp] += 1

    b2b_ano = [e for e in b2b if (e.get("cadastro") or "").startswith(ano_str)]
    b2c_ano = [e for e in b2c if (e.get("cadastro") or "").startswith(ano_str)]
    enr_ano = [e for e in enrollments if (e.get("cadastro") or "").startswith(ano_str)]

    if ano == ano_atual:
        mes_anterior = (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
        vendas_kpi = len([e for e in enr_ano if (e.get("cadastro") or "").startswith(mes_atual)])
        vendas_ant = len([e for e in enrollments if (e.get("cadastro") or "").startswith(mes_anterior)])
        kpi_label = "Vendas este mês"
        kpi_compare = "vs mês passado"
    else:
        ano_ant_str = str(ano - 1)
        vendas_kpi = len(enr_ano)
        vendas_ant = len([e for e in enrollments if (e.get("cadastro") or "").startswith(ano_ant_str)])
        kpi_label = f"Vendas em {ano}"
        kpi_compare = f"vs {ano - 1}"

    var_pct = round((vendas_kpi - vendas_ant) / vendas_ant * 100, 1) if vendas_ant else 0

    empresa_count_ano = defaultdict(int)
    for e in b2b_ano:
        emp = extract_empresa(e.get("grupo_nome", ""))
        if emp:
            empresa_count_ano[emp] += 1
    top10_clientes = sorted(empresa_count_ano.items(), key=lambda x: x[1], reverse=True)[:10]

    curso_count = defaultdict(int)
    for e in enr_ano:
        curso_count[e.get("titulo_curso") or "Sem nome"] += 1
    top10_cursos = sorted(curso_count.items(), key=lambda x: x[1], reverse=True)[:10]

    trend = {f"{ano_str}-{m:02d}": {"b2b": 0, "b2c": 0} for m in range(1, 13)}
    for e in b2b_ano:
        mes = (e.get("cadastro") or "")[:7]
        if mes in trend:
            trend[mes]["b2b"] += 1
    for e in b2c_ano:
        mes = (e.get("cadastro") or "")[:7]
        if mes in trend:
            trend[mes]["b2c"] += 1

    return {
        "resumo": {
            "total_matriculas_ativas": total_active,
            "b2b_ativas": len(b2b_active),
            "b2c_ativas": len(b2c_active),
            "pct_b2b": round(len(b2b_active) / total_active * 100, 1) if total_active else 0,
            "pct_b2c": round(len(b2c_active) / total_active * 100, 1) if total_active else 0,
            "vendas_kpi": vendas_kpi,
            "vendas_ant": vendas_ant,
            "var_pct": var_pct,
            "kpi_label": kpi_label,
            "kpi_compare": kpi_compare,
            "total_empresas_b2b": len(empresa_count_ativas),
            "mes_atual": mes_atual,
            "ano": ano,
        },
        "top10_clientes": [{"empresa": k, "matriculas": v} for k, v in top10_clientes],
        "top10_cursos": [{"curso": k, "matriculas": v} for k, v in top10_cursos],
        "trend_mensal": [
            {"mes": k, "b2b": v["b2b"], "b2c": v["b2c"], "atual": k == mes_atual}
            for k, v in sorted(trend.items())
        ],
    }


def compute_funil(progress_data):
    counts = {"Concluído": 0, "Em Andamento": 0, "Não Iniciado": 0}
    for p in progress_data:
        s = p.get("situacao") or "Não Iniciado"
        if s in counts:
            counts[s] += 1
        else:
            counts["Não Iniciado"] += 1
    total = sum(counts.values())
    return {
        "concluido": counts["Concluído"],
        "andamento": counts["Em Andamento"],
        "nao_iniciado": counts["Não Iniciado"],
        "total": total,
    }


def compute_novos_recorrentes(enrollments, ano):
    ano_str = str(ano)
    primeiro_ano = {}
    for e in enrollments:
        aid = e.get("aluno_id")
        ano_e = (e.get("cadastro") or "")[:4]
        if aid and ano_e.isdigit():
            a = int(ano_e)
            if aid not in primeiro_ano or a < primeiro_ano[aid]:
                primeiro_ano[aid] = a

    alunos_no_ano = set()
    for e in enrollments:
        if (e.get("cadastro") or "").startswith(ano_str):
            aid = e.get("aluno_id")
            if aid:
                alunos_no_ano.add(aid)

    novos = sum(1 for aid in alunos_no_ano if primeiro_ano.get(aid) == ano)
    recorrentes = len(alunos_no_ano) - novos
    total = len(alunos_no_ano)
    return {
        "novos": novos,
        "recorrentes": recorrentes,
        "total": total,
        "pct_novos": round(novos / total * 100, 1) if total else 0,
        "pct_recorrentes": round(recorrentes / total * 100, 1) if total else 0,
    }


def compute_conclusao_cursos(enrollments, certificates, ano, min_enrolled=5):
    ano_str = str(ano)
    enrolled = defaultdict(set)
    for e in enrollments:
        if (e.get("cadastro") or "").startswith(ano_str):
            titulo = e.get("titulo_curso") or "Sem nome"
            aid = e.get("aluno_id")
            if aid:
                enrolled[titulo].add(aid)

    certified = defaultdict(set)
    for c in certificates:
        titulo = c.get("curso_titulo") or "Sem nome"
        aid = c.get("aluno_id")
        if aid:
            certified[titulo].add(aid)

    result = []
    for titulo, alunos in enrolled.items():
        if len(alunos) < min_enrolled:
            continue
        cert_count = len(certified.get(titulo, set()) & alunos)
        taxa = round(cert_count / len(alunos) * 100, 1)
        result.append({
            "curso": titulo,
            "matriculados": len(alunos),
            "certificados": cert_count,
            "taxa": taxa,
        })

    result.sort(key=lambda x: x["matriculados"], reverse=True)
    return result[:15]


CURSOS_EXCLUIDOS_RECERT = ["NR-35", "NR 35"]


def _curso_excluido(titulo):
    titulo_upper = (titulo or "").upper()
    return any(excl.upper() in titulo_upper for excl in CURSOS_EXCLUIDOS_RECERT)


def norma_base(titulo):
    if not titulo:
        return ""
    t = titulo.strip()
    t = re.sub(r'\s*-\s*\d+\s*(anos?|h|horas?)\b.*$', '', t, flags=re.IGNORECASE)
    t = re.sub(r'\b(19|20)\d{2}\b', '', t)
    t = re.sub(r'\bNR\s+(\d+)\b', r'NR-\1', t, flags=re.IGNORECASE)
    t = re.sub(r'[\s\-]+$', '', t)
    t = re.sub(r'\s{2,}', ' ', t)
    return t.strip()


def compute_recert_lista(certificates, students_map, emails, canais, empresas):
    now = datetime.now()
    regras = get_validades()
    ultimo = {}
    for c in certificates:
        aluno = c.get("aluno_id")
        titulo = c.get("curso_titulo") or ""
        concl = (c.get("concluido") or "")[:10]
        if aluno is None or not concl:
            continue
        if _curso_excluido(titulo):
            continue
        chave = (aluno, norma_base(titulo))
        atual = ultimo.get(chave)
        if atual is None or concl > (atual.get("concluido") or "")[:10]:
            ultimo[chave] = c

    lista = []
    for c in ultimo.values():
        concl = c.get("concluido")
        titulo = c.get("curso_titulo") or ""
        try:
            dt = datetime.strptime(str(concl)[:10], "%Y-%m-%d")
            dt_recert = dt + timedelta(days=validade_curso(titulo, regras))
            dias = (dt_recert - now).days
            if 0 <= dias <= 90:
                aid = c.get("aluno_id")
                s = students_map.get(aid, {})
                lista.append({
                    "aluno_id": aid,
                    "curso_id": c.get("curso_id"),
                    "aluno": c.get("aluno_nome", ""),
                    "email": emails.get(aid, ""),
                    "telefone": s.get("telefone", ""),
                    "cpf": s.get("cpf", ""),
                    "cidade": s.get("cidade", ""),
                    "uf": normalizar_uf(s.get("uf", "")),
                    "canal": canais.get(aid, "B2C"),
                    "empresa": empresas.get(aid, "") if canais.get(aid) == "B2B" else "",
                    "curso": titulo,
                    "nota": c.get("media_final", ""),
                    "concluido": concl[:10] if concl else "",
                    "recertifica": dt_recert.strftime("%Y-%m-%d"),
                    "dias": dias,
                    "ultimo_acesso": (s.get("ultimo_acesso") or "")[:10],
                    "pdf": c.get("certificado_pdf", ""),
                })
        except (ValueError, TypeError, KeyError):
            pass
    lista.sort(key=lambda x: x["dias"])
    return lista


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/dashboard")
def dashboard():
    if count_table("enrollments") == 0:
        return jsonify({"error": "sem_cache"}), 503

    enrollments = load_table("enrollments")
    certificates = load_table("certificates")
    students_map = mapa_student()
    emails = mapa_email()
    canais = mapa_canal()
    empresas = mapa_empresa()

    ano_atual = datetime.now().year
    try:
        ano = int(request.args.get("ano", ano_atual))
    except (ValueError, TypeError):
        ano = ano_atual

    data = compute_vendas(enrollments, ano)
    lista = compute_recert_lista(certificates, students_map, emails, canais, empresas)

    r30 = len([x for x in lista if x["dias"] <= 30])
    r60 = len([x for x in lista if 30 < x["dias"] <= 60])
    r90 = len([x for x in lista if 60 < x["dias"] <= 90])

    geo = compute_geo(enrollments, students_map)

    data["recert"] = {"r30": r30, "r60": r60, "r90": r90, "total": len(lista)}
    data["geo"] = [{"uf": k, "total": v} for k, v in geo[:15]]
    data["funil"] = {"concluido": 0, "andamento": 0, "nao_iniciado": 0, "total": 0}
    data["novos_recorrentes"] = compute_novos_recorrentes(enrollments, ano)
    data["conclusao_cursos"] = compute_conclusao_cursos(enrollments, certificates, ano)
    data["cache_info"] = {
        "total_matriculas": count_table("enrollments"),
        "total_certificados": count_table("certificates"),
        "total_alunos": count_table("students"),
        "ultimo_sync": get_meta("last_sync"),
        "ultimo_sync_modo": get_meta("last_sync_modo") or "completo",
        "anos": anos_disponiveis(enrollments),
    }
    return jsonify(data)


@app.route("/api/exportar")
def exportar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    certificates = load_table("certificates")
    students_map = mapa_student()
    emails = mapa_email()
    canais = mapa_canal()
    empresas = mapa_empresa()
    lista = compute_recert_lista(certificates, students_map, emails, canais, empresas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Recompra - Recertificacao"

    headers = [
        "Prazo (dias)", "Faixa", "Aluno", "E-mail", "Telefone", "CPF",
        "Cidade", "UF", "Canal", "Empresa", "Curso", "Nota Final",
        "Concluído em", "Recertifica em", "Último Acesso", "Link do Certificado",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1A1D27", end_color="1A1D27", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    fill_30 = PatternFill(start_color="FDE7E7", end_color="FDE7E7", fill_type="solid")
    fill_60 = PatternFill(start_color="FEF3E2", end_color="FEF3E2", fill_type="solid")
    fill_90 = PatternFill(start_color="E7F0FD", end_color="E7F0FD", fill_type="solid")
    cols_center = {1, 2, 8, 9, 12, 13, 14, 15}

    for item in lista:
        dias = item["dias"]
        faixa = "Urgente (0-30)" if dias <= 30 else ("Atenção (31-60)" if dias <= 60 else "Planejamento (61-90)")
        fill = fill_30 if dias <= 30 else (fill_60 if dias <= 60 else fill_90)
        row_data = [
            dias, faixa, item["aluno"], item["email"], item["telefone"], item["cpf"],
            item["cidade"], item["uf"], item["canal"], item["empresa"], item["curso"],
            item["nota"], item["concluido"], item["recertifica"], item["ultimo_acesso"], item["pdf"],
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            if col in cols_center:
                cell.alignment = Alignment(horizontal="center")
            if col == 2:
                cell.fill = fill

    widths = [12, 18, 30, 32, 16, 15, 18, 6, 8, 30, 45, 10, 14, 14, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"recompra_recertificacao_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/sync", methods=["POST"])
def sync():
    if sync_status["running"]:
        return jsonify({"ok": False, "msg": "Já em andamento."})
    modo = (request.get_json(silent=True) or {}).get("modo", "rapido")
    if modo not in ("rapido", "completo"):
        modo = "rapido"
    t = threading.Thread(target=do_sync, args=(modo,), daemon=True)
    t.start()
    return jsonify({"ok": True, "modo": modo})


@app.route("/api/sync/status")
def sync_status_route():
    return jsonify(sync_status)


EMAIL_REMETENTE = os.getenv("EMAIL_REMETENTE", "bosstreinamentos@hotmail.com")
EMAIL_SENHA = os.getenv("EMAIL_SENHA", "")
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587

email_status = {"running": False, "total": 0, "enviados": 0, "pulados": 0, "erros": [], "msg": "", "concluido": False}


def _faixa_label(dias):
    if dias <= 30:
        return 30
    if dias <= 60:
        return 60
    return 90


def _ja_enviou(conn, aluno_id, curso_id, faixa):
    limite = (datetime.now() - timedelta(days=25)).isoformat()
    c = conn.cursor()
    c.execute(
        "SELECT id FROM emails_enviados WHERE aluno_id=%s AND curso_id=%s AND faixa=%s AND data_envio>%s",
        (aluno_id, curso_id, faixa, limite)
    )
    row = c.fetchone()
    c.close()
    return row is not None


def _registrar_envio(conn, aluno_id, curso_id, faixa):
    c = conn.cursor()
    c.execute(
        "INSERT INTO emails_enviados (aluno_id, curso_id, faixa, data_envio) VALUES (%s,%s,%s,%s)",
        (aluno_id, curso_id, faixa, datetime.now().isoformat())
    )
    conn.commit()
    c.close()


def _corpo_email(nome, curso, dias, empresa=None):
    urgencia = "em breve" if dias > 30 else f"em apenas {dias} dias"
    saudacao = nome.split()[0].capitalize() if nome else "Aluno"
    rodape_empresa = f"<p style='color:#888;font-size:13px'>Empresa: <b>{empresa}</b></p>" if empresa else ""
    return f"""
<html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:20px">
<div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1)">
  <div style="background:#1a1d27;padding:28px 32px;text-align:center">
    <span style="color:#f5a623;font-weight:900;font-size:28px;letter-spacing:2px">BOSS</span><br>
    <span style="color:#888;font-size:12px;letter-spacing:1px">CONSULTORIA E TREINAMENTOS</span>
  </div>
  <div style="padding:32px">
    <h2 style="color:#1a1d27;margin-bottom:8px">Olá, {saudacao}!</h2>
    <p style="color:#444;line-height:1.6">
      Seu certificado do curso <strong>{curso}</strong> vence <strong>{urgencia}</strong>.
    </p>
    <p style="color:#444;line-height:1.6;margin-top:12px">
      Para manter sua conformidade legal, é necessário realizar a <strong>reciclagem</strong>
      antes do vencimento. A Boss Treinamentos já tem turmas disponíveis — entre em contato
      e garanta sua vaga com antecedência.
    </p>
    <div style="text-align:center;margin:28px 0">
      <a href="https://bosstreinamentos.com"
         style="background:#f5a623;color:#000;text-decoration:none;padding:14px 32px;border-radius:8px;font-weight:700;font-size:15px">
        Ver cursos disponíveis
      </a>
    </div>
    {rodape_empresa}
    <hr style="border:none;border-top:1px solid #eee;margin:24px 0">
    <p style="color:#aaa;font-size:12px;text-align:center">
      Boss Consultoria e Treinamentos · <a href="https://bosstreinamentos.com" style="color:#f5a623">bosstreinamentos.com</a><br>
      Para cancelar o recebimento destes avisos, entre em contato pelo site.
    </p>
  </div>
</div>
</body></html>"""


def _do_envio(lista, faixa_filtro):
    global email_status
    conn = get_conn()
    try:
        smtp = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
        smtp.starttls()
        smtp.login(EMAIL_REMETENTE, EMAIL_SENHA)

        emails_map = mapa_email()

        for item in lista:
            dias = item.get("dias", 999)
            if dias > 90:
                continue

            faixa = _faixa_label(dias)
            if faixa_filtro and faixa != faixa_filtro:
                continue

            aluno_id = item.get("aluno_id")
            curso_id = item.get("curso_id")
            email_dest = item.get("email") or emails_map.get(aluno_id)

            if not email_dest or not aluno_id or not curso_id:
                email_status["pulados"] += 1
                continue

            if _ja_enviou(conn, aluno_id, curso_id, faixa):
                email_status["pulados"] += 1
                continue

            try:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = f"Certificado vencendo em {dias} dias — {item.get('curso', '')}"
                msg["From"] = f"Boss Treinamentos <{EMAIL_REMETENTE}>"
                msg["To"] = email_dest
                msg["Cc"] = "bosstreinamentos@hotmail.com"
                msg.attach(MIMEText(_corpo_email(item.get("aluno", ""), item.get("curso", ""), dias, item.get("empresa")), "html", "utf-8"))
                smtp.sendmail(EMAIL_REMETENTE, [email_dest, "bosstreinamentos@hotmail.com"], msg.as_string())
                _registrar_envio(conn, aluno_id, curso_id, faixa)
                email_status["enviados"] += 1
                email_status["msg"] = f"{email_status['enviados']} de {email_status['total']} enviados"
            except Exception as e:
                email_status["erros"].append(f"{email_dest}: {type(e).__name__}")
                app.logger.exception("Erro ao enviar e-mail para %s", email_dest)

        smtp.quit()
        email_status["msg"] = f"Concluído — {email_status['enviados']} enviados, {email_status['pulados']} pulados"
    except smtplib.SMTPAuthenticationError:
        email_status["msg"] = "Erro: falha de autenticação — verifique EMAIL_SENHA no .env"
        app.logger.error("SMTPAuthenticationError ao enviar e-mails")
    except Exception as e:
        email_status["msg"] = f"Erro de conexão SMTP: {type(e).__name__}"
        app.logger.exception("Erro SMTP geral")
    finally:
        conn.close()
        email_status["running"] = False
        email_status["concluido"] = True


@app.route("/api/enviar_recert/teste", methods=["POST"])
def enviar_recert_teste():
    if not EMAIL_SENHA:
        return jsonify({"erro": "EMAIL_SENHA não configurada no .env"}), 500

    email_dest = (request.json or {}).get("email", "").strip()
    if not email_dest:
        return jsonify({"erro": "Informe o campo 'email' no body da requisição"}), 400

    try:
        smtp = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
        smtp.starttls()
        smtp.login(EMAIL_REMETENTE, EMAIL_SENHA)

        msg = MIMEMultipart("alternative")
        msg["Subject"] = "[TESTE] Certificado vencendo em 15 dias — NR-35 Trabalho em Altura"
        msg["From"] = f"Boss Treinamentos <{EMAIL_REMETENTE}>"
        msg["To"] = email_dest

        corpo = _corpo_email(
            nome="Fulano da Silva",
            curso="NR-35 - Trabalho em Altura 2025",
            dias=15,
            empresa="Empresa Exemplo Ltda"
        )
        msg.attach(MIMEText(corpo, "html", "utf-8"))
        smtp.sendmail(EMAIL_REMETENTE, email_dest, msg.as_string())
        smtp.quit()

        return jsonify({"ok": True, "enviado_para": email_dest})
    except smtplib.SMTPAuthenticationError:
        return jsonify({"erro": "Falha de autenticação — verifique EMAIL_SENHA no .env"}), 401
    except Exception as e:
        app.logger.exception("Erro no envio de teste")
        return jsonify({"erro": f"{type(e).__name__}: {e}"}), 500


@app.route("/api/enviar_recert", methods=["POST"])
def enviar_recert():
    global email_status
    if email_status["running"]:
        return jsonify({"erro": "Envio já em andamento"}), 409
    if not EMAIL_SENHA:
        return jsonify({"erro": "EMAIL_SENHA não configurada no .env"}), 500

    faixa_filtro = None
    if request.json and request.json.get("faixa"):
        faixa_filtro = int(request.json["faixa"])

    certificates = load_table("certificates")
    students_map = mapa_student()
    emails_map = mapa_email()
    lista = compute_recert_lista(certificates, students_map, emails_map, mapa_canal(), mapa_empresa())
    candidatos = [x for x in lista if x.get("dias", 999) <= 90]
    limite = int((request.json or {}).get("limite", 0))
    if limite > 0:
        candidatos = candidatos[:limite]

    email_status = {
        "running": True,
        "total": len(candidatos),
        "enviados": 0,
        "pulados": 0,
        "erros": [],
        "msg": f"Iniciando envio para {len(candidatos)} candidatos...",
        "concluido": False,
    }

    t = threading.Thread(target=_do_envio, args=(candidatos, faixa_filtro), daemon=True)
    t.start()
    return jsonify({"ok": True, "total": len(candidatos)})


@app.route("/api/email/status")
def email_status_route():
    return jsonify(email_status)


@app.route("/api/email/stats")
def email_stats():
    certificates = load_table("certificates")
    students_map = mapa_student()
    emails_map = mapa_email()
    lista = compute_recert_lista(certificates, students_map, emails_map, mapa_canal(), mapa_empresa())
    total_fila = len(lista)

    if total_fila == 0:
        return jsonify({"total_fila": 0, "ja_notificados": 0, "pendentes": 0, "ultimo_envio": None})

    conn = get_conn()
    c = conn.cursor()
    limite = (datetime.now() - timedelta(days=25)).isoformat()
    ja_notificados = 0
    for item in lista:
        faixa = _faixa_label(item["dias"])
        c.execute(
            "SELECT id FROM emails_enviados WHERE aluno_id=%s AND curso_id=%s AND faixa=%s AND data_envio>%s",
            (item.get("aluno_id"), item.get("curso_id"), faixa, limite)
        )
        if c.fetchone():
            ja_notificados += 1

    c.execute("SELECT data_envio FROM emails_enviados ORDER BY data_envio DESC LIMIT 1")
    ultimo = c.fetchone()
    c.close()
    conn.close()

    return jsonify({
        "total_fila": total_fila,
        "ja_notificados": ja_notificados,
        "pendentes": total_fila - ja_notificados,
        "ultimo_envio": ultimo[0][:10] if ultimo else None,
    })


@app.route("/api/validades", methods=["GET"])
def api_validades_get():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, padrao, dias, ordem FROM validades ORDER BY ordem")
    rows = c.fetchall()
    c.close()
    conn.close()
    return jsonify([{"id": r[0], "padrao": r[1], "dias": r[2], "ordem": r[3]} for r in rows])


@app.route("/api/validades", methods=["POST"])
def api_validades_post():
    items = request.get_json(force=True) or []
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM validades")
    for i, item in enumerate(items):
        padrao = str(item.get("padrao", "")).strip().lower()
        dias = int(item.get("dias", 730))
        if padrao:
            c.execute("INSERT INTO validades (padrao, dias, ordem) VALUES (%s,%s,%s)", (padrao, dias, i))
    conn.commit()
    c.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/colaboradores/upload", methods=["POST"])
def colaboradores_upload():
    from openpyxl import load_workbook

    if "arquivo" not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    f = request.files["arquivo"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"erro": "Envie um arquivo .xlsx ou .xls."}), 400

    def norm_cpf(v):
        return re.sub(r"[.\-\s/]", "", str(v).strip())

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"erro": f"Erro ao ler planilha: {e}"}), 400

    if not rows:
        return jsonify({"erro": "Planilha vazia."}), 400

    header = [str(c or "").lower().strip() for c in rows[0]]
    col_cpf, col_email = None, None
    for i, h in enumerate(header):
        if col_cpf is None and any(k in h for k in ("cpf", "documento", "doc")):
            col_cpf = i
        if col_email is None and any(k in h for k in ("email", "e-mail", "correio")):
            col_email = i

    termos_email = set()
    termos_cpf = set()
    data_rows = rows[1:] if (col_cpf is not None or col_email is not None) else rows

    for row in data_rows:
        if col_email is not None and col_email < len(row):
            v = str(row[col_email] or "").strip()
            if "@" in v:
                termos_email.add(v.lower())
        if col_cpf is not None and col_cpf < len(row):
            v = norm_cpf(row[col_cpf] or "")
            if len(v) >= 11:
                termos_cpf.add(v)
        if col_cpf is None and col_email is None:
            for cell in row:
                v = str(cell or "").strip()
                if "@" in v:
                    termos_email.add(v.lower())
                else:
                    n = norm_cpf(v)
                    if len(n) == 11 and n.isdigit():
                        termos_cpf.add(n)

    if not termos_email and not termos_cpf:
        return jsonify({"erro": "Nenhum CPF ou e-mail encontrado na planilha."}), 400

    students = load_table("students")
    certificates = load_table("certificates")

    certs_por_aluno = defaultdict(list)
    for c in certificates:
        aid = c.get("aluno_id")
        if aid and c.get("certificado_pdf") and c.get("concluido"):
            certs_por_aluno[aid].append(c)

    encontrados = {}
    for s in students:
        aid = s.get("aluno_id")
        email_s = (s.get("email") or "").lower()
        cpf_s = norm_cpf(s.get("cpf") or "")
        if email_s in termos_email or (len(cpf_s) >= 11 and cpf_s in termos_cpf):
            if aid not in encontrados:
                certs = certs_por_aluno.get(aid, [])
                encontrados[aid] = {
                    "aluno_id": aid,
                    "nome": s.get("nome", ""),
                    "email": s.get("email", ""),
                    "cpf": s.get("cpf", ""),
                    "certificados": [
                        {
                            "curso_id": c.get("curso_id"),
                            "titulo": c.get("curso_titulo", ""),
                            "concluido": (c.get("concluido") or "")[:10],
                            "pdf": c.get("certificado_pdf", ""),
                        }
                        for c in sorted(certs, key=lambda x: x.get("concluido") or "", reverse=True)
                    ],
                }

    return jsonify({
        "alunos": list(encontrados.values()),
        "lidos": len(termos_email) + len(termos_cpf),
    })


@app.route("/api/colaboradores/zip", methods=["POST"])
def colaboradores_zip():
    import zipfile
    body = request.get_json(force=True) or {}
    itens = body.get("itens", [])
    if not itens:
        return jsonify({"erro": "Nenhum certificado selecionado."}), 400

    buf = io.BytesIO()
    erros = []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in itens:
            pdf_url = item.get("pdf", "")
            nome_aluno = re.sub(r"[^\w\s\-]", "", item.get("nome", "Aluno")).strip()
            titulo = re.sub(r"[^\w\s\-]", "", item.get("titulo", "Curso")).strip()
            concluido = item.get("concluido", "")
            filename = f"{nome_aluno} — {titulo} ({concluido}).pdf"
            try:
                r = requests.get(pdf_url, timeout=20, headers=HEADERS)
                if r.status_code == 200:
                    zf.writestr(filename, r.content)
                else:
                    erros.append(f"{nome_aluno}/{titulo}: HTTP {r.status_code}")
            except Exception as e:
                erros.append(f"{nome_aluno}/{titulo}: {e}")

    buf.seek(0)
    if erros:
        app.logger.warning("ZIP — erros ao baixar PDFs: %s", erros)

    return send_file(
        buf,
        as_attachment=True,
        download_name="certificados.zip",
        mimetype="application/zip",
    )


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("⚠  Falta a biblioteca openpyxl. Rode: pip install openpyxl")
    print("=" * 50)
    print("Boss Dashboard — http://localhost:5000")
    print(f"Matrículas: {count_table('enrollments')} | Certificados: {count_table('certificates')} | Alunos: {count_table('students')} | Progresso: {count_table('progress')}")
    print("=" * 50)
    app.run(debug=True, port=5000)
